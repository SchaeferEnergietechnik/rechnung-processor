#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Rechnungsbearbeitung - PDF-Rechnungen verarbeiten (Python Version)

Funktionen:
- Liest PDFs aus dem Ordner, in dem das Script liegt
- Extrahiert: Gesamtpreis (inkl. MwSt), Lieferant, Rechnungsdatum, Rechnungsnummer
- OCR-Unterstützung für gescannte PDFs (falls Tesseract installiert)
- Benennt um: XX-xxx_Lieferant_YYYY-MM-DD_Gesamtpreis.pdf
- Jährliche fortlaufende Nummerierung (25-001, 25-002... 26-001, 26-002)
- Duplikat-Check via Rechnungsnummer
- Schreibt Excel mit Auswertung, deutschem Datum und Monatssummen

Usage: python rechnung_processor.py
    oder als .exe mit PyInstaller
"""

import os
import re
import sys
import tempfile
import subprocess
import json
from pathlib import Path
from datetime import datetime
from io import BytesIO
from collections import defaultdict

try:
    import requests
    REQUESTS_VERFUEGBAR = True
except ImportError:
    REQUESTS_VERFUEGBAR = False
    print("⚠️  requests nicht installiert. Online-OCR deaktiviert.")
    print("   Installieren mit: pip install requests")

try:
    import pdfplumber
except ImportError:
    print("Fehler: pdfplumber nicht installiert.")
    print("Installieren mit: pip install pdfplumber openpyxl")
    sys.exit(1)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("Fehler: openpyxl nicht installiert.")
    print("Installieren mit: pip install openpyxl")
    sys.exit(1)

# Optional: OCR-Unterstützung
try:
    from pdf2image import convert_from_path
    import pytesseract
    OCR_VERFUEGBAR = True
except ImportError:
    OCR_VERFUEGBAR = False

# Konfiguration
CONFIG = {
    "excel_datei": "Rechnungsuebersicht.xlsx",
    "ocr_space_api_key": "K83767756488957"  # Free OCR API Key
}

def get_prefix_for_year(jahr):
    """Generiert Präfix aus dem Jahr (z.B. 2025 -> 25-)"""
    return str(jahr)[-2:] + "-"

# Bekannte Lieferanten mit korrekten Umlauten - SPEZIFISCHE zuerst!
LIEFERANTEN_MAP = {
    # Spezifische Treffer zuerst (längere/genaue Strings)
    're-invent retail': 'RE-INvent Retail',
    'voelkner': 'Voelkner',
    'steinke': 'Steinke',
    'ttcher': 'Böttcher',
    'bottcher': 'Böttcher',
    'stadtwerke wittenberge': 'Stadtwerke Wittenberge',
    'amazon': 'Amazon',
    # Allgemeine Treffer zuletzt
    'stadtwerke': 'Stadtwerke',
    'wittenberge': 'Stadtwerke Wittenberge',
    'invent': 'RE-INvent',
}

# Stoppwörter - ERWEITERT
STOPP_WORTER = ['umsatzsteuer', 'ust-id', 'steuernummer', 'bankverbindung', 'bic', 'iban', 
    'öffnungszeiten', 'montag', 'dienstag', 'mittwoch', 'donnerstag', 'freitag', 
    'strasse', 'ferchlipp', 'lichterfelde', 'altmärkische', 'deutschland', 'g.e.s.', 'energietechnik',
    'gesamtbetrag', 'rechnungsbetrag', 'endbetrag', 'brutto', 'netto', 'zzgl', 'zuzüglich',
    'abzüglich', 'abschlag', 'zahlung', 'betrag', 'summe', 'bruttowert', 'nettowert']


def parse_betrag(betrag_str):
    """Betrag normalisieren (String zu Float) - unterstützt Punkt und Komma"""
    if not betrag_str:
        return 0.0
    # Verschiedene Formate: 1.234,56 oder 1,234.56 oder 1234.56
    betrag_str = str(betrag_str).strip()
    
    # Zähle Kommas und Punkte
    comma_count = betrag_str.count(',')
    dot_count = betrag_str.count('.')
    
    # Deutsch: 1.234,56 (Punkt = Tausender, Komma = Dezimal)
    # Englisch: 1,234.56 (Komma = Tausender, Punkt = Dezimal)
    
    if comma_count == 1 and dot_count == 0:
        # Nur Komma -> Deutsch: 1234,56
        clean = betrag_str.replace(',', '.')
    elif dot_count == 1 and comma_count == 0:
        # Nur Punkt -> Englisch: 1234.56
        clean = betrag_str
    elif comma_count >= 1 and dot_count >= 1:
        # Beide vorhanden - letztes ist Dezimaltrenner
        last_comma = betrag_str.rfind(',')
        last_dot = betrag_str.rfind('.')
        if last_comma > last_dot:
            # Komma ist Dezimal: 1.234,56
            clean = betrag_str.replace('.', '').replace(',', '.')
        else:
            # Punkt ist Dezimal: 1,234.56
            clean = betrag_str.replace(',', '')
    else:
        clean = betrag_str
    
    try:
        return float(clean)
    except ValueError:
        return 0.0


def format_betrag(betrag):
    """Betrag formatieren für Dateinamen (deutsch mit Komma)"""
    return f"{betrag:.2f}".replace('.', ',')


def finde_lieferant(text):
    """Extrahiert den Lieferanten aus dem Text"""
    text_lower = text.lower()
    lines = text.split('\n')
    
    # HARTE REGEL: Bekannte Lieferanten direkt erkennen
    if 'stadtwerke wittenberge gmbh' in text_lower:
        return 'Stadtwerke_Wittenberge'
    if 'böttcher ag' in text_lower:
        return 'Böttcher'
    if 'bottcher ag' in text_lower:
        return 'Böttcher'
    if 're-invent retail' in text_lower or 're-invent' in text_lower:
        return 'RE-INvent_Retail'
    if 'voelkner' in text_lower:
        return 'Voelkner'
    if 'steinke' in text_lower:
        return 'Steinke'
    if 'ttt-filmservice' in text_lower or 'filmservice' in text_lower:
        return 'TTT-Filmservice'
    
    # AMAZON SPEZIAL
    if 'amazon eu' in text_lower or 'amazon' in text_lower:
        if 'bestellnummer' in text_lower or 'amazon.de' in text_lower or 'verkauft von' in text_lower:
            return 'Amazon'
    
    # Suche nach explizitem "Rechnung" + Firma
    rechnung_pattern = re.search(r'ReCHNUNG\s*\n?\s*([A-Z][A-Za-zäöüß\s&\.]+(?:GmbH|AG|KG|OHG|e\.?K|UG))', text, re.IGNORECASE)
    if rechnung_pattern:
        kandidat = rechnung_pattern.group(1).strip()
        if ist_gueltiger_lieferant(kandidat):
            return bereinige_lieferant(kandidat)
    
    # Suche nach Firmen die G.E.S. beliefert
    ges_pattern = re.search(r'([A-Z][A-Za-zäöüß\s&\.]+(?:GmbH|AG|KG|OHG|e\.?K|UG))[^.]{0,200}G\.E\.S', text, re.IGNORECASE | re.DOTALL)
    if ges_pattern:
        kandidat = ges_pattern.group(1).strip()
        if ist_gueltiger_lieferant(kandidat):
            return bereinige_lieferant(kandidat)
    
    # Erste 10 Zeilen = meist Absender
    absender_section = '\n'.join(lines[:10])
    firmen_absender = re.findall(r'([A-Z][A-Za-zäöüß\s&\.]+(?:GmbH|AG|KG|OHG|e\.?K|UG))', absender_section)
    for firma in firmen_absender:
        firma_clean = firma.strip()
        if ist_gueltiger_lieferant(firma_clean):
            return bereinige_lieferant(firma_clean)
    
    # Fallback: Erste Zeilen durchgehen
    for zeile in lines[:15]:
        if re.search(r'\b(GmbH|AG|KG|OHG|e\.?K|UG)\b', zeile, re.IGNORECASE):
            bereinigt = bereinige_lieferant(zeile)
            if ist_gueltiger_lieferant(bereinigt):
                return bereinigt
    
    return 'Unbekannt'


def ist_gueltiger_lieferant(name):
    """Prüft ob ein Name ein gültiger Lieferant ist (kein Stoppwort)"""
    if len(name) < 5:
        return False
    name_lower = name.lower()
    for stop in STOPP_WORTER:
        if stop in name_lower:
            return False
    if re.match(r'^(gesamt|rechnungs|end|brutto|netto|summe)', name_lower):
        return False
    return True


def bereinige_lieferant(name):
    """Bereinigt den Lieferantennamen für Dateinamen"""
    fixed = name.strip()
    fixed = re.sub(r'bottcher', 'Böttcher', fixed, flags=re.IGNORECASE)
    fixed = re.sub(r'schaefer', 'Schäfer', fixed, flags=re.IGNORECASE)
    
    if 'ä' not in fixed and 'ö' not in fixed and 'ü' not in fixed:
        fixed = fixed.replace('ae', 'ä').replace('oe', 'ö').replace('ue', 'ü').replace('ss', 'ß')
    
    fixed = re.sub(r'[\\/:*?"<>|]', '_', fixed)
    fixed = re.sub(r'\s+', '_', fixed)
    fixed = re.sub(r'_+', '_', fixed)
    fixed = fixed.strip('_')
    
    return fixed[:40]


def finde_rechnungsnummer(text):
    """Extrahiert die Rechnungsnummer aus dem Text"""
    patterns = [
        r'Rechnungsnummer\s*:?\s*([A-Z0-9\-]+)',
        r'Rechnungsnummer[:\s]+(\d{4,20})',
        r'(?:Rechnung|Rechnungsnr)[\.\s#:]*([A-Z0-9\-]{3,30})',
        r'(?:R|RG)[\-\s]?[Nn]r?\.?\s*:?\s*([A-Z0-9\-]{3,30})',
    ]
    
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        if matches:
            best_match = max(matches, key=len)
            best_match = best_match.strip().strip(':')
            if len(best_match) >= 3:
                return best_match
    
    return None


def finde_gesamtpreis(text):
    """Findet den BRUTTO-Gesamtbetrag (inkl. MwSt)"""
    text_lower = text.lower()
    alle_kandidaten = []
    
    # Flexible Muster für verschiedene Zahlenformate
    hoch_prio_patterns = [
        (r'(?:summe|gesamtpreis|zahlbetrag)[\s:]+(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})\s*(?:eur|€)?', 110),
        (r'(?:summe|gesamtpreis)[\s:]+(?:eur|€)?\s*(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})', 110),
        (r'zahlbetrag[\s:]+(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})', 105),
        (r'(?:noch\s+zu\s+zahlender\s+betrag|rechnungsbetrag)[^\d]{0,50}(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})', 100),
        (r'(?:zu\s*zahlen|fällig)[^\d]{0,50}(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})', 90),
        (r'zahlbar[^\d]{0,100}(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})', 80),
        (r'gesamtpreis\s+(?:eur|€)?\s*(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})', 110),
        (r'endbetrag[\s:]+(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})', 100),
        (r'summe\s+(?:bruttowert)?[\s:]+(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})', 95),
    ]
    
    for pattern, prio in hoch_prio_patterns:
        matches = list(re.finditer(pattern, text, re.IGNORECASE))
        for m in matches:
            betrag = parse_betrag(m.group(1))
            if 0 < betrag < 50000:
                context_start = max(0, m.start() - 30)
                context = text[context_start:m.start()].lower()
                if any(bad in context for bad in ['versandkosten', 'rabatt', 'discount', 'einzelpreis', 'zwischensumme']):
                    continue
                alle_kandidaten.append((betrag, prio, m.group(1), "endbetrag"))
    
    # Bruttowert explizit
    match = re.search(r'summe\s+bruttowert[:\s]*(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})', text, re.IGNORECASE)
    if match:
        betrag = parse_betrag(match.group(1))
        if 0 < betrag < 50000:
            alle_kandidaten.append((betrag, 95, match.group(1), "bruttowert"))
    
    # Gesamtbetrag
    gesamt_matches = list(re.finditer(r'gesamtbetrag[:\s]*(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})', text, re.IGNORECASE))
    if len(gesamt_matches) > 1:
        betraege = [(parse_betrag(m.group(1)), m.group(1)) for m in gesamt_matches if 0 < parse_betrag(m.group(1)) < 50000]
        if betraege:
            max_betrag, original = max(betraege, key=lambda x: x[0])
            alle_kandidaten.append((max_betrag, 70, original, "gesamtbetrag_max"))
    elif gesamt_matches:
        betrag = parse_betrag(gesamt_matches[0].group(1))
        if 0 < betrag < 50000:
            alle_kandidaten.append((betrag, 70, gesamt_matches[0].group(1), "gesamtbetrag"))
    
    # Fallback: EUR-Beträge in der Nähe von "gesamt" oder "summe"
    if not alle_kandidaten:
        eur_matches = list(re.finditer(r'(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})\s*(?:eur|€)', text, re.IGNORECASE))
        for m in eur_matches:
            context_start = max(0, m.start() - 50)
            context = text[context_start:m.start()].lower()
            if any(good in context for good in ['gesamt', 'summe', 'betrag', 'endbetrag']):
                betrag = parse_betrag(m.group(1))
                if 0 < betrag < 50000:
                    alle_kandidaten.append((betrag, 60, m.group(1), "eur_kontext"))
    
    if alle_kandidaten:
        alle_kandidaten.sort(key=lambda x: (x[1], x[0]), reverse=True)
        best_betrag, best_prio, best_orig, best_typ = alle_kandidaten[0]
        return {"betrag": best_betrag, "original": best_orig, "typ": best_typ}
    
    return None


def finde_datum(text):
    """Extrahiert das Rechnungsdatum aus dem Text"""
    text_clean = text.replace('\n', ' ')
    
    rechnungs_patterns = [
        r'ReCHNUNGSDATUM[:\s]+(\d{1,2})[\.\/](\d{1,2})[\.\/](\d{4})',
        r'ReCHNUNG[^\n]{0,100}DATUM[:\s]+(\d{1,2})[\.\/](\d{1,2})[\.\/](\d{4})',
        r'DATUM[:\s]+(\d{1,2})[\.\/](\d{1,2})[\.\/](\d{4})[^\n]{0,50}RECHNUNG',
    ]
    
    for pattern in rechnungs_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            tag = int(match.group(1))
            monat = int(match.group(2))
            jahr = int(match.group(3))
            if 1 <= tag <= 31 and 1 <= monat <= 12 and 2020 <= jahr <= 2030:
                iso = f"{jahr:04d}-{monat:02d}-{tag:02d}"
                return {"tag": tag, "monat": monat, "jahr": jahr, "iso": iso}
    
    datum_pattern = r'(?:\D|^)(\d{1,2})[\.\/](\d{1,2})[\.\/](\d{4})(?:\D|$)'
    alle_datum = []
    
    for match in re.finditer(datum_pattern, text_clean):
        tag = int(match.group(1))
        monat = int(match.group(2))
        jahr = int(match.group(3))
        
        if 1 <= tag <= 31 and 1 <= monat <= 12 and 2020 <= jahr <= 2030:
            context_start = max(0, match.start() - 100)
            context_end = min(len(text_clean), match.end() + 50)
            context = text_clean[context_start:context_end].lower()
            
            ist_rechnung = re.search(r'rechnung|rechnungsdatum|fällig|zahlbar', context, re.IGNORECASE)
            ist_lieferung = re.search(r'liefer|versand', context, re.IGNORECASE)
            ist_bestellung = re.search(r'bestell', context, re.IGNORECASE)
            
            gewicht = 3 if ist_rechnung else (2 if ist_lieferung else (1 if ist_bestellung else 0))
            if gewicht > 0:
                iso = f"{jahr:04d}-{monat:02d}-{tag:02d}"
                alle_datum.append({"tag": tag, "monat": monat, "jahr": jahr, "iso": iso, "gewicht": gewicht})
    
    if not alle_datum:
        return None
    
    beste_kandidaten = [(d["gewicht"], d["jahr"], d["monat"], d["tag"], d["iso"]) for d in alle_datum]
    beste_kandidaten.sort(reverse=True)
    _, jahr, monat, tag, iso = beste_kandidaten[0]
    return {"tag": tag, "monat": monat, "jahr": jahr, "iso": iso}


def ocr_pdf_online(datei_pfad):
    """Verwendet OCR.space API für gescannte PDFs"""
    if not REQUESTS_VERFUEGBAR:
        return None
    
    api_key = CONFIG.get("ocr_space_api_key", "")
    if not api_key:
        return None
    
    datei_groesse = os.path.getsize(datei_pfad)
    if datei_groesse > 1024 * 1024:
        print(f"  ⚠️  Datei zu groß für Free-OCR ({datei_groesse / 1024 / 1024:.2f} MB > 1 MB)")
        return None
    
    try:
        print("  🌐 Versuche Online-OCR (OCR.space)...")
        url = "https://api.ocr.space/parse/image"
        
        with open(datei_pfad, 'rb') as f:
            response = requests.post(
                url,
                files={"file": (os.path.basename(datei_pfad), f)},
                data={
                    "apikey": api_key,
                    "language": "ger",
                    "OCREngine": "2",
                    "isTable": "false",
                    "detectOrientation": "true",
                    "scale": "true"
                },
                timeout=120
            )
        
        if response.status_code == 200:
            result = response.json()
            
            if result.get("OCRExitCode") == 1:
                parsed_text = ""
                for parsed_result in result.get("ParsedResults", []):
                    parsed_text += parsed_result.get("ParsedText", "") + "\n"
                
                if parsed_text.strip():
                    print("  ✅ Online-OCR erfolgreich")
                    return parsed_text
                else:
                    print("  ⚠️  Online-OCR: Kein Text erkannt")
                    return None
            else:
                error_msg = result.get("ErrorMessage", ["Unbekannter Fehler"])[0]
                print(f"  ⚠️  Online-OCR Fehler: {error_msg}")
                return None
        else:
            print(f"  ⚠️  Online-OCR HTTP Fehler: {response.status_code}")
            return None
            
    except requests.exceptions.Timeout:
        print("  ⚠️  Online-OCR Timeout (120s)")
        return None
    except Exception as e:
        print(f"  ⚠️  Online-OCR Fehler: {e}")
        return None


def ocr_pdf(datei_pfad):
    """Versucht OCR auf gescanntem PDF durchzuführen"""
    if not OCR_VERFUEGBAR:
        return None
    
    try:
        bilder = convert_from_path(datei_pfad, dpi=300, first_page=1, last_page=3)
        
        text = ""
        for bild in bilder:
            seiten_text = pytesseract.image_to_string(bild, lang='deu')
            text += seiten_text + "\n"
        
        return text if len(text.strip()) > 50 else None
    except Exception as e:
        print(f"  OCR-Fehler: {e}")
        return None


def verarbeite_pdf(datei_pfad):
    """Verarbeitet eine einzelne PDF-Datei"""
    try:
        text = ""
        seiten = 0
        ist_gescannt = False
        
        with pdfplumber.open(datei_pfad) as pdf:
            seiten = len(pdf.pages)
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        
        if len(text.strip()) < 100:
            ist_gescannt = True
            print("  📄 Wenig Text gefunden, versuche OCR...")
            
            ocr_text = ocr_pdf(datei_pfad)
            if ocr_text:
                text = ocr_text
                ist_gescannt = False
                print("  ✅ Lokales OCR erfolgreich (Tesseract)")
            else:
                ocr_text = ocr_pdf_online(datei_pfad)
                if ocr_text:
                    text = ocr_text
                    ist_gescannt = True
                    print("  ✅ Online-OCR erfolgreich (OCR.space)")
                else:
                    if OCR_VERFUEGBAR or REQUESTS_VERFUEGBAR:
                        return {
                            "datei": os.path.basename(datei_pfad),
                            "pfad": datei_pfad,
                            "gescannt": True,
                            "extrahiert": False,
                            "hinweis": "Gescanntes PDF - OCR fehlgeschlagen"
                        }
                    else:
                        return {
                            "datei": os.path.basename(datei_pfad),
                            "pfad": datei_pfad,
                            "gescannt": True,
                            "extrahiert": False,
                            "hinweis": "Gescanntes PDF - keine OCR verfügbar"
                        }
        
        lieferant = finde_lieferant(text)
        gesamtpreis = finde_gesamtpreis(text)
        datum = finde_datum(text)
        rechnungsnummer = finde_rechnungsnummer(text)
        
        return {
            "datei": os.path.basename(datei_pfad),
            "pfad": datei_pfad,
            "lieferant": lieferant,
            "gesamtpreis": gesamtpreis["betrag"] if gesamtpreis else 0.0,
            "gesamtpreis_original": gesamtpreis["original"] if gesamtpreis else '-',
            "gesamtpreis_typ": gesamtpreis["typ"] if gesamtpreis else '-',
            "datum": datum,
            "rechnungsnummer": rechnungsnummer if rechnungsnummer else '-',
            "seiten": seiten,
            "ocr_verwendet": ist_gescannt,
            "extrahiert": True
        }
        
    except Exception as e:
        return {
            "datei": os.path.basename(datei_pfad),
            "pfad": datei_pfad,
            "fehler": str(e),
            "extrahiert": False
        }


def generiere_dateiname(daten, nummer, ist_duplikat=False, prefix="26-"):
    """Generiert neuen Dateinamen"""
    nummer_str = f"{prefix}{nummer:03d}"
    lieferant_safe = daten["lieferant"]
    datum_str = daten["datum"]["iso"] if daten["datum"] else "0000-00-00"
    preis_str = format_betrag(daten["gesamtpreis"])
    
    if ist_duplikat:
        return f"doppelt_{nummer_str}_{lieferant_safe}_{datum_str}_{preis_str}EUR.pdf"
    return f"{nummer_str}_{lieferant_safe}_{datum_str}_{preis_str}EUR.pdf"


def lade_excel_bestand(ziel_pfad):
    """Lädt bestehende Excel und extrahiert Rechnungsnummern für Duplikat-Check"""
    bestehende_eintraege = []
    bekannte_rechnungsnummern = set()
    jahres_hoechste_nummer = defaultdict(int)
    
    if not os.path.exists(ziel_pfad):
        return bestehende_eintraege, bekannte_rechnungsnummern, jahres_hoechste_nummer
    
    try:
        wb_alt = load_workbook(ziel_pfad)
        ws_alt = wb_alt["Rechnungen"]
        
        for row in ws_alt.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:  # Lfd. Nr. und Rechnungsnummer
                lfd_nr = str(row[0])
                rechnungsnr = str(row[1]) if row[1] else ""
                datum_iso = str(row[2]) if row[2] else "-"
                
                bestehende_eintraege.append({
                    "lfd_nr": lfd_nr,
                    "rechnungsnummer": rechnungsnr,
                    "datum_iso": datum_iso,
                    "datum_de": str(row[3]) if len(row) > 3 else "-",
                    "lieferant": str(row[4]) if len(row) > 4 else "",
                    "gesamtpreis": str(row[5]) if len(row) > 5 else "",
                    "seiten": row[6] if len(row) > 6 else 0,
                    "original": str(row[7]) if len(row) > 7 else "",
                    "neu_name": str(row[8]) if len(row) > 8 else ""
                })
                
                # Rechnungsnummer speichern für Duplikat-Check
                if rechnungsnr and rechnungsnr != '-':
                    bekannte_rechnungsnummern.add(rechnungsnr)
                
                # Jährliche höchste Nummer ermitteln (z.B. "25-001" -> Jahr 2025, Nummer 1)
                if datum_iso != "-" and len(datum_iso) >= 4:
                    jahr = datum_iso[:4]
                    try:
                        match = re.search(r'(\d{2})-(\d{3})', lfd_nr)
                        if match:
                            nr = int(match.group(2))
                            jahres_hoechste_nummer[jahr] = max(jahres_hoechste_nummer[jahr], nr)
                    except:
                        pass
        
        print(f"📊 {len(bestehende_eintraege)} bestehende Einträge geladen")
        print(f"   {len(bekannte_rechnungsnummern)} bekannte Rechnungsnummern")
        
    except Exception as e:
        print(f"⚠️  Konnte bestehende Excel nicht laden: {e}")
    
    return bestehende_eintraege, bekannte_rechnungsnummern, jahres_hoechste_nummer


def erstelle_excel(ergebnisse, ziel_pfad, quell_ordner, bestehende_eintraege, bekannte_rechnungsnummern):
    """Aktualisiert oder erstellt die Excel-Datei"""
    
    neue_eintraege = []
    duplikate = []
    
    for e in ergebnisse:
        if not e["extrahiert"]:
            continue
        
        rechnungsnr = e.get("rechnungsnummer", "-")
        datum_iso = e["datum"]["iso"] if e["datum"] else None
        
        # Duplikat-Check
        if rechnungsnr != "-" and rechnungsnr in bekannte_rechnungsnummern:
            duplikate.append(e)
            continue
        
        # Nummer aus dem bereits generierten Dateinamen extrahieren
        neuer_name = e.get("neuer_name", "")
        lfd_nr = e.get("lfd_nr", "")
        
        if not lfd_nr and neuer_name:
            match = re.match(r'(?:doppelt_)?(\d{2}-\d{3})_', neuer_name)
            if match:
                lfd_nr = match.group(1)
        
        if not lfd_nr:
            lfd_nr = "00-000"
        
        # Deutsch Format
        if datum_iso:
            parts = datum_iso.split("-")
            datum_de = f"{parts[2]}.{parts[1]}.{parts[0]}"
        else:
            datum_de = "-"
        
        neue_eintraege.append({
            "lfd_nr": lfd_nr,
            "rechnungsnummer": rechnungsnr,
            "datum_iso": datum_iso if datum_iso else "-",
            "datum_de": datum_de,
            "lieferant": e["lieferant"],
            "gesamtpreis": format_betrag(e["gesamtpreis"]),
            "seiten": e["seiten"],
            "original": e["datei"],
            "neu_name": neuer_name,
            "neu_name_duplikat": f"doppelt_{neuer_name}" if not neuer_name.startswith("doppelt_") else neuer_name
        })
        
        # Rechnungsnummer zum Bestand hinzufügen
        if rechnungsnr != "-":
            bekannte_rechnungsnummern.add(rechnungsnr)
    
    # Alle Einträge kombinieren
    alle_eintraege = bestehende_eintraege + neue_eintraege
    
    # Sortiere nach Datum
    def sort_key(e):
        datum = e["datum_iso"] if e["datum_iso"] != "-" else "0000-00-00"
        return datum
    
    alle_eintraege.sort(key=sort_key)
    
    # Gruppiere nach Jahr-Monat für Summen
    monats_gruppen = defaultdict(list)
    for e in alle_eintraege:
        if e["datum_iso"] != "-":
            key = e["datum_iso"][:7]
            monats_gruppen[key].append(e)
        else:
            monats_gruppen["0000-00"].append(e)
    
    # Workbook erstellen
    wb = Workbook()
    ws_haupt = wb.active
    ws_haupt.title = "Rechnungen"
    
    # Header
    headers = ["Lfd. Nr.", "Rechnungsnummer", "Rechnungsdatum (ISO)", "Rechnungsdatum", 
               "Lieferant", "Gesamtpreis inkl. MwSt (€)", "Anzahl Seiten", 
               "Ursprünglicher Dateiname", "Neuer Dateiname"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_haupt.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # Daten schreiben
    current_row = 2
    monats_summen = {}
    
    for monat_key in sorted(monats_gruppen.keys()):
        eintraege = monats_gruppen[monat_key]
        monats_summe = 0.0
        
        for e in eintraege:
            ws_haupt.cell(row=current_row, column=1, value=e["lfd_nr"])
            ws_haupt.cell(row=current_row, column=2, value=e["rechnungsnummer"])
            ws_haupt.cell(row=current_row, column=3, value=e["datum_iso"])
            ws_haupt.cell(row=current_row, column=4, value=e["datum_de"])
            ws_haupt.cell(row=current_row, column=5, value=e["lieferant"])
            ws_haupt.cell(row=current_row, column=6, value=e["gesamtpreis"])
            ws_haupt.cell(row=current_row, column=7, value=e["seiten"])
            ws_haupt.cell(row=current_row, column=8, value=e["original"])
            ws_haupt.cell(row=current_row, column=9, value=e["neu_name"])
            
            try:
                betrag_str = str(e["gesamtpreis"]).replace(".", "").replace(",", ".")
                monats_summe += float(betrag_str)
            except:
                pass
            
            current_row += 1
        
        # Monatssumme
        monats_name = monat_key
        if monat_key != "0000-00":
            parts = monat_key.split("-")
            monats_name = f"{parts[1]}.{parts[0]}"
        
        sum_cell = ws_haupt.cell(row=current_row, column=1, value=f"Summe {monats_name}")
        sum_cell.font = Font(bold=True)
        sum_cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        sum_value = ws_haupt.cell(row=current_row, column=6, value=format_betrag(monats_summe))
        sum_value.font = Font(bold=True)
        sum_value.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        monats_summen[monat_key] = monats_summe
        current_row += 1
    
    # Spaltenbreiten
    ws_haupt.column_dimensions['A'].width = 12
    ws_haupt.column_dimensions['B'].width = 25
    ws_haupt.column_dimensions['C'].width = 20
    ws_haupt.column_dimensions['D'].width = 18
    ws_haupt.column_dimensions['E'].width = 30
    ws_haupt.column_dimensions['F'].width = 25
    ws_haupt.column_dimensions['G'].width = 15
    ws_haupt.column_dimensions['H'].width = 45
    ws_haupt.column_dimensions['I'].width = 50
    
    # ISO-Spalte verstecken
    ws_haupt.column_dimensions['C'].hidden = True
    
    # Monatsauswertung
    ws_monat = wb.create_sheet("Monatsauswertung")
    for col, header in enumerate(["Jahr-Monat", "Monatsname", "Anzahl Rechnungen", "Gesamtsumme (€)"], 1):
        cell = ws_monat.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    for idx, (key, summe) in enumerate(sorted(monats_summen.items()), 1):
        if key != "0000-00":
            parts = key.split("-")
            jahr = parts[0]
            monat = int(parts[1])
            monatsname = datetime(int(jahr), monat, 1).strftime("%B %Y")
            anzahl = len(monats_gruppen[key])
            
            ws_monat.cell(row=idx + 1, column=1, value=key)
            ws_monat.cell(row=idx + 1, column=2, value=monatsname)
            ws_monat.cell(row=idx + 1, column=3, value=anzahl)
            ws_monat.cell(row=idx + 1, column=4, value=format_betrag(summe))
    
    ws_monat.column_dimensions['A'].width = 15
    ws_monat.column_dimensions['B'].width = 20
    ws_monat.column_dimensions['C'].width = 20
    ws_monat.column_dimensions['D'].width = 20
    
    # Jahresauswertung
    ws_jahr = wb.create_sheet("Jahresauswertung")
    jahres_daten = defaultdict(lambda: {"anzahl": 0, "summe": 0.0})
    
    for key, summe in monats_summen.items():
        if key != "0000-00":
            jahr = key.split("-")[0]
            jahres_daten[jahr]["anzahl"] += len(monats_gruppen[key])
            jahres_daten[jahr]["summe"] += summe
    
    for col, header in enumerate(["Jahr", "Anzahl Rechnungen", "Gesamtsumme (€)"], 1):
        cell = ws_jahr.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    for idx, (jahr, daten) in enumerate(sorted(jahres_daten.items()), 1):
        ws_jahr.cell(row=idx + 1, column=1, value=jahr)
        ws_jahr.cell(row=idx + 1, column=2, value=daten["anzahl"])
        ws_jahr.cell(row=idx + 1, column=3, value=format_betrag(daten["summe"]))
    
    ws_jahr.column_dimensions['A'].width = 15
    ws_jahr.column_dimensions['B'].width = 20
    ws_jahr.column_dimensions['C'].width = 20
    
    wb.save(ziel_pfad)
    
    return len(neue_eintraege), len(alle_eintraege), duplikate


def main():
    """Hauptfunktion"""
    quell_ordner = os.getcwd()
    
    print(f"📁 Verarbeite PDFs in: {quell_ordner}\n")
    
    if OCR_VERFUEGBAR:
        print("✅ OCR-Unterstützung verfügbar\n")
    else:
        print("ℹ️  OCR nicht verfügbar - nur digitale PDFs\n")
    
    # Lade bestehende Excel
    excel_pfad = os.path.join(quell_ordner, CONFIG["excel_datei"])
    bestehende_eintraege, bekannte_rechnungsnummern, jahres_hoechste_nummer = lade_excel_bestand(excel_pfad)
    
    # Finde alle PDFs (ohne bereits umbenannte)
    dateien = [
        os.path.join(quell_ordner, f) for f in os.listdir(quell_ordner)
        if f.lower().endswith('.pdf') 
        and not re.match(r'\d{2}-\d{3}_', f)  # XX-XXX_ Format
        and not f.lower().startswith('doppelt_')
    ]
    
    if not dateien:
        print("⚠️  Keine neuen PDF-Dateien gefunden.")
        return
    
    print(f"📄 {len(dateien)} neue PDF(s) gefunden. Verarbeite...\n")
    
    # Verarbeite alle PDFs
    ergebnisse = []
    jahres_counter = defaultdict(int)  # Zählt neue Nummern pro Jahr
    
    for i, pfad in enumerate(dateien, 1):
        datei_name = os.path.basename(pfad)
        print(f"[{i}/{len(dateien)}] {datei_name}...")
        
        daten = verarbeite_pdf(pfad)
        
        if daten["extrahiert"]:
            rechnungsnr = daten.get("rechnungsnummer", "-")
            datum_iso = daten["datum"]["iso"] if daten["datum"] else None
            
            # Duplikat-Check
            if rechnungsnr != "-" and rechnungsnr in bekannte_rechnungsnummern:
                daten["ist_duplikat"] = True
                prefix = get_prefix_for_year(daten["datum"]["jahr"]) if daten["datum"] else "00-"
                daten["neuer_name"] = generiere_dateiname(daten, 0, ist_duplikat=True, prefix=prefix)
                print(f"  ⚠️  Duplikat (Rechnungsnr: {rechnungsnr})")
            else:
                # Jahr ermitteln
                if datum_iso:
                    jahr_key = datum_iso[:4]
                    jahr_int = int(jahr_key)
                else:
                    jahr_key = "0000"
                    jahr_int = datetime.now().year
                
                # Nächste Nummer = höchste aus Bestand + laufende + 1
                hoechste = jahres_hoechste_nummer.get(jahr_key, 0)
                aktuelle_nummer = hoechste + jahres_counter[jahr_key] + 1
                jahres_counter[jahr_key] += 1
                
                # Präfix vom Rechnungsjahr ableiten
                prefix = get_prefix_for_year(jahr_int)
                
                daten["ist_duplikat"] = False
                daten["neuer_name"] = generiere_dateiname(daten, aktuelle_nummer, prefix=prefix)
                daten["lfd_nr"] = f"{prefix}{aktuelle_nummer:03d}"
                
                if rechnungsnr != "-":
                    bekannte_rechnungsnummern.add(rechnungsnr)
            
            # Umbenennen
            ziel_pfad_datei = os.path.join(quell_ordner, daten["neuer_name"])
            try:
                os.rename(pfad, ziel_pfad_datei)
                print(f"  ✅ {daten['neuer_name']}")
            except Exception as e:
                print(f"  ⚠️  Umbenennen fehlgeschlagen: {e}")
            
            print(f"     Lieferant: {daten['lieferant']}")
            print(f"     Rechnungsnr: {daten['rechnungsnummer']}")
            print(f"     Datum: {daten['datum']['iso'] if daten['datum'] else '-'}")
            print(f"     Preis: {format_betrag(daten['gesamtpreis'])} €")
            if daten.get("ocr_verwendet"):
                print(f"     ℹ️  OCR verwendet")
        else:
            print(f"  ⚠️  {daten.get('hinweis', daten.get('fehler', 'Fehler'))}")
        
        ergebnisse.append(daten)
        print()
    
    # Excel erstellen
    neue_eintraege, gesamt_eintraege, duplikate = erstelle_excel(
        ergebnisse, excel_pfad, quell_ordner, 
        bestehende_eintraege, bekannte_rechnungsnummern
    )
    
    # Zusammenfassung
    erfolgreich = len([e for e in ergebnisse if e["extrahiert"]])
    gesamt_summe = sum(e["gesamtpreis"] for e in ergebnisse if e["extrahiert"] and not e.get("ist_duplikat", False))
    
    print("\n📊 Zusammenfassung:")
    print(f"   Verarbeitet: {erfolgreich}")
    print(f"   Neue Einträge: {neue_eintraege}")
    print(f"   Duplikate: {len(duplikate)}")
    print(f"   Gesamtsumme (neu): {format_betrag(gesamt_summe)} €")
    print(f"\n📂 Excel: {CONFIG['excel_datei']} ({gesamt_eintraege} Einträge)")
    
    print("\nDrücken Sie Enter zum Beenden...")
    input()


if __name__ == "__main__":
    main()
