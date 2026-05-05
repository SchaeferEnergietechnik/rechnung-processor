#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Excel-Auffüller - Trägt bereits umbenannte PDFs ins Excel ein

Usage: python excel_auffueller.py
"""

import os
import re
from datetime import datetime
from collections import defaultdict

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("Fehler: openpyxl nicht installiert.")
    print("Installieren mit: pip install openpyxl")
    exit(1)

CONFIG = {
    "excel_datei": "Rechnungsuebersicht.xlsx"
}


def parse_betrag(betrag_str):
    """Betrag normalisieren (String zu Float)"""
    if not betrag_str:
        return 0.0
    betrag_str = str(betrag_str).strip()
    
    comma_count = betrag_str.count(',')
    dot_count = betrag_str.count('.')
    
    if comma_count == 1 and dot_count == 0:
        clean = betrag_str.replace(',', '.')
    elif dot_count == 1 and comma_count == 0:
        clean = betrag_str
    elif comma_count >= 1 and dot_count >= 1:
        last_comma = betrag_str.rfind(',')
        last_dot = betrag_str.rfind('.')
        if last_comma > last_dot:
            clean = betrag_str.replace('.', '').replace(',', '.')
        else:
            clean = betrag_str.replace(',', '')
    else:
        clean = betrag_str
    
    try:
        return float(clean)
    except ValueError:
        return 0.0


def format_betrag(betrag):
    """Betrag formatieren für Excel (deutsch mit Komma)"""
    return f"{betrag:.2f}".replace('.', ',')


def lade_excel_bestand(ziel_pfad):
    """Lädt bestehende Excel"""
    bestehende_eintraege = []
    bekannte_rechnungsnummern = set()
    
    if not os.path.exists(ziel_pfad):
        return bestehende_eintraege, bekannte_rechnungsnummern
    
    try:
        wb_alt = load_workbook(ziel_pfad)
        ws_alt = wb_alt["Rechnungen"]
        
        for row in ws_alt.iter_rows(min_row=2, values_only=True):
            if row[0]:
                lfd_nr = str(row[0])
                rechnungsnr = str(row[1]) if row[1] else ""
                datum_iso = str(row[2]) if row[2] else "-"
                
                bestehende_eintraege.append({
                    "lfd_nr": lfd_nr,
                    "rechnungsnummer": rechnungsnr,
                    "datum_iso": datum_iso,
                    "datum_de": str(row[3]) if len(row) > 3 else "-",
                    "lieferant": str(row[4]) if len(row) > 4 else "",
                    "gesamtpreis": str(row[5]) if len(row) > 5 else "",
                    "seiten": row[6] if len(row) > 6 else 0,
                    "original": str(row[7]) if len(row) > 7 else "",
                    "neu_name": str(row[8]) if len(row) > 8 else ""
                })
                
                if rechnungsnr and rechnungsnr != '-':
                    bekannte_rechnungsnummern.add(rechnungsnr)
        
        print(f"📊 {len(bestehende_eintraege)} bestehende Einträge geladen")
        
    except Exception as e:
        print(f"⚠️  Konnte bestehende Excel nicht laden: {e}")
    
    return bestehende_eintraege, bekannte_rechnungsnummern


def erstelle_excel(alle_eintraege, ziel_pfad):
    """Erstellt die Excel-Datei"""
    
    # Sortiere nach Datum
    def sort_key(e):
        datum = e["datum_iso"] if e["datum_iso"] != "-" else "0000-00-00"
        return datum
    
    alle_eintraege.sort(key=sort_key)
    
    # Gruppiere nach Jahr-Monat
    monats_gruppen = defaultdict(list)
    for e in alle_eintraege:
        if e["datum_iso"] != "-":
            key = e["datum_iso"][:7]
            monats_gruppen[key].append(e)
        else:
            monats_gruppen["0000-00"].append(e)
    
    # Workbook erstellen
    wb = Workbook()
    ws_haupt = wb.active
    ws_haupt.title = "Rechnungen"
    
    # Header
    headers = ["Lfd. Nr.", "Rechnungsnummer", "Rechnungsdatum (ISO)", "Rechnungsdatum", 
               "Lieferant", "Gesamtpreis inkl. MwSt (€)", "Anzahl Seiten", 
               "Ursprünglicher Dateiname", "Neuer Dateiname"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_haupt.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # Daten schreiben
    current_row = 2
    monats_summen = {}
    
    for monat_key in sorted(monats_gruppen.keys()):
        eintraege = monats_gruppen[monat_key]
        monats_summe = 0.0
        
        for e in eintraege:
            ws_haupt.cell(row=current_row, column=1, value=e["lfd_nr"])
            ws_haupt.cell(row=current_row, column=2, value=e["rechnungsnummer"])
            ws_haupt.cell(row=current_row, column=3, value=e["datum_iso"])
            ws_haupt.cell(row=current_row, column=4, value=e["datum_de"])
            ws_haupt.cell(row=current_row, column=5, value=e["lieferant"])
            ws_haupt.cell(row=current_row, column=6, value=e["gesamtpreis"])
            ws_haupt.cell(row=current_row, column=7, value=e["seiten"])
            ws_haupt.cell(row=current_row, column=8, value=e["original"])
            ws_haupt.cell(row=current_row, column=9, value=e["neu_name"])
            
            try:
                betrag_str = str(e["gesamtpreis"]).replace(".", "").replace(",", ".")
                monats_summe += float(betrag_str)
            except:
                pass
            
            current_row += 1
        
        # Monatssumme
        monats_name = monat_key
        if monat_key != "0000-00":
            parts = monat_key.split("-")
            monats_name = f"{parts[1]}.{parts[0]}"
        
        sum_cell = ws_haupt.cell(row=current_row, column=1, value=f"Summe {monats_name}")
        sum_cell.font = Font(bold=True)
        sum_cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        sum_value = ws_haupt.cell(row=current_row, column=6, value=format_betrag(monats_summe))
        sum_value.font = Font(bold=True)
        sum_value.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        monats_summen[monat_key] = monats_summe
        current_row += 1
    
    # Spaltenbreiten
    ws_haupt.column_dimensions['A'].width = 12
    ws_haupt.column_dimensions['B'].width = 25
    ws_haupt.column_dimensions['C'].width = 20
    ws_haupt.column_dimensions['D'].width = 18
    ws_haupt.column_dimensions['E'].width = 30
    ws_haupt.column_dimensions['F'].width = 25
    ws_haupt.column_dimensions['G'].width = 15
    ws_haupt.column_dimensions['H'].width = 45
    ws_haupt.column_dimensions['I'].width = 50
    ws_haupt.column_dimensions['C'].hidden = True
    
    # Monatsauswertung
    ws_monat = wb.create_sheet("Monatsauswertung")
    for col, header in enumerate(["Jahr-Monat", "Monatsname", "Anzahl Rechnungen", "Gesamtsumme (€)"], 1):
        cell = ws_monat.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    for idx, (key, summe) in enumerate(sorted(monats_summen.items()), 1):
        if key != "0000-00":
            parts = key.split("-")
            jahr = parts[0]
            monat = int(parts[1])
            monatsname = datetime(int(jahr), monat, 1).strftime("%B %Y")
            anzahl = len(monats_gruppen[key])
            
            ws_monat.cell(row=idx + 1, column=1, value=key)
            ws_monat.cell(row=idx + 1, column=2, value=monatsname)
            ws_monat.cell(row=idx + 1, column=3, value=anzahl)
            ws_monat.cell(row=idx + 1, column=4, value=format_betrag(summe))
    
    ws_monat.column_dimensions['A'].width = 15
    ws_monat.column_dimensions['B'].width = 20
    ws_monat.column_dimensions['C'].width = 20
    ws_monat.column_dimensions['D'].width = 20
    
    wb.save(ziel_pfad)
    print(f"\n💾 Excel gespeichert: {ziel_pfad}")


def main():
    """Hauptfunktion - parst bereits umbenannte PDFs"""
    ordner = os.getcwd()
    
    print(f"📁 Suche nach umbenannten PDFs in: {ordner}\n")
    
    # Lade bestehende Excel
    excel_pfad = os.path.join(ordner, CONFIG["excel_datei"])
    bestehende_eintraege, _ = lade_excel_bestand(excel_pfad)
    
    # Suche nach bereits umbenannten PDFs (Format: XX-XXX_Lieferant_Datum_PreisEUR.pdf)
    pattern = r'^(\d{2}-\d{3})_(.+)_(\d{4}-\d{2}-\d{2})_(.+?)EUR\.pdf$'
    
    dateien = [
        f for f in os.listdir(ordner)
        if f.lower().endswith('.pdf') and re.match(pattern, f, re.IGNORECASE)
    ]
    
    if not dateien:
        print("⚠️  Keine umbenannten PDFs gefunden.")
        print("   Erwartetes Format: XX-XXX_Lieferant_YYYY-MM-DD_PreisEUR.pdf")
        return
    
    print(f"📄 {len(dateien)} umbenannte PDF(s) gefunden\n")
    
    neue_eintraege = []
    
    for datei_name in dateien:
        match = re.match(pattern, datei_name, re.IGNORECASE)
        if match:
            lfd_nr = match.group(1)
            lieferant = match.group(2).replace('_', ' ')
            datum_iso = match.group(3)
            preis_str = match.group(4).replace('_', ',')
            
            # Deutsch Format
            parts = datum_iso.split("-")
            datum_de = f"{parts[2]}.{parts[1]}.{parts[0]}"
            
            neue_eintraege.append({
                "lfd_nr": lfd_nr,
                "rechnungsnummer": "-",  # Aus Dateiname nicht ermittelbar
                "datum_iso": datum_iso,
                "datum_de": datum_de,
                "lieferant": lieferant,
                "gesamtpreis": preis_str.replace('.', ','),
                "seiten": 0,  # Unbekannt
                "original": datei_name,
                "neu_name": datei_name
            })
            
            print(f"✅ {datei_name}")
            print(f"   Lfd.Nr: {lfd_nr}, Lieferant: {lieferant}, Datum: {datum_de}, Preis: {preis_str} €")
    
    # Kombinieren und speichern
    alle_eintraege = bestehende_eintraege + neue_eintraege
    
    if neue_eintraege:
        erstelle_excel(alle_eintraege, excel_pfad)
        print(f"\n📊 {len(neue_eintraege)} neue Einträge hinzugefügt")
        print(f"   {len(alle_eintraege)} Einträge insgesamt")
    else:
        print("\n⚠️  Keine neuen Einträge zu speichern")
    
    print("\nDrücken Sie Enter zum Beenden...")
    input()


if __name__ == "__main__":
    main()
